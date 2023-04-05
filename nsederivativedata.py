import requests
from openpyxl import Workbook,load_workbook



baseurl = "https://www.nseindia.com/"
url = f"https://www.nseindia.com/api/liveEquity-derivatives?index=top20_contracts"
headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, '
                         'like Gecko) '
                         'Chrome/80.0.3987.149 Safari/537.36',
           'accept-language': 'en,gu;q=0.9,hi;q=0.8', 'accept-encoding': 'gzip, deflate, br'}
session = requests.Session()
request = session.get(baseurl, headers=headers, timeout=5)
cookies = dict(request.cookies)
response = session.get(url, headers=headers, timeout=5, cookies=cookies)

#x is a dictionary that has data
x = response.json()
timestamp=x["timestamp"]
marketStatus=x["marketStatus"]
arr=[]
for elm in x["data"]:
    arr.append(elm)



print(timestamp)
print(marketStatus)
# Creating an excel file

wb = load_workbook("./Derivatives.xlsx")
ws = wb.active
i=1


ws[str(chr(66))+str(i)].value='instrument'
ws[str(chr(67))+str(i)].value='expiryDate'
ws[str(chr(68))+str(i)].value='optionType'
ws[str(chr(69))+str(i)].value='strikePrice'
ws[str(chr(70))+str(i)].value='lastPrice'
ws[str(chr(71))+str(i)].value='change'
ws[str(chr(72))+str(i)].value='pChange'
ws[str(chr(73))+str(i)].value='volume'
ws[str(chr(74))+str(i)].value='value'
ws[str(chr(75))+str(i)].value='openInterest'
ws[str(chr(76))+str(i)].value='underlyingValue'

for elm in arr:
    i=i+1
    ws[str(chr(66))+str(i)].value=elm['instrument']
    ws[str(chr(67))+str(i)].value=elm['expiryDate']
    ws[str(chr(68))+str(i)].value=elm['optionType']
    ws[str(chr(69))+str(i)].value=elm['strikePrice']
    ws[str(chr(70))+str(i)].value=elm['lastPrice']
    ws[str(chr(71))+str(i)].value=elm['change']
    ws[str(chr(72))+str(i)].value=elm['pChange']
    ws[str(chr(73))+str(i)].value=elm['volume']
    ws[str(chr(74))+str(i)].value=elm['value']
    ws[str(chr(75))+str(i)].value=elm['openInterest']
    ws[str(chr(76))+str(i)].value=elm['underlyingValue']
    
wb.save("./Derivatives.xlsx")
